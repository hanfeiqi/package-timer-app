import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide", page_title="📦 Package SLA Analyzer")

st.title("📦 Package SLA Delivery Analyzer")

# ========== 文件上传 ==========
uploaded_file = st.file_uploader("Upload CSV file", type=["csv"])
if not uploaded_file:
    st.warning("Please upload a CSV file to proceed.")
    st.stop()

# ========== 读取数据 ==========
df = pd.read_csv(uploaded_file)

# ========== 字段转换 ==========
time_fields = ["GOFO签入时间", "目的中心签入时间", "最新领件时间", "妥投时间", "统计日期"]
for col in time_fields:
    df[col] = pd.to_datetime(df[col], errors="coerce")

# ========== 耗时计算 ==========
df['GOFO_to_中心_hrs'] = (df['目的中心签入时间'] - df['GOFO签入时间']).dt.total_seconds() / 3600
df['中心_to_领件_hrs'] = (df['最新领件时间'] - df['目的中心签入时间']).dt.total_seconds() / 3600
df['领件_to_妥投_hrs'] = (df['妥投时间'] - df['最新领件时间']).dt.total_seconds() / 3600
df['Total_Duration_Hrs'] = df[['GOFO_to_中心_hrs', '中心_to_领件_hrs', '领件_to_妥投_hrs']].sum(axis=1)

# ========== 包裹状态 ==========
def classify_status(row):
    if pd.notna(row['妥投时间']):
        return '已妥投'
    elif pd.notna(row['最新领件时间']):
        return '已领件未妥投'
    elif pd.notna(row['目的中心签入时间']):
        return '已签入未领件'
    else:
        return '未签入'
df['包裹状态'] = df.apply(classify_status, axis=1)

# ========== 时效设置 ==========
time_limits = {
    'SAN': 48, 'LAV': 48, 'VTC': 48, 'CNO': 48, 'BKD': 48, 'SFV': 48,
    'SLC': 96, 'STG': 96, 'TUC': 96, 'SFO': 96, 'SMF': 96, 'RNO': 96, 'RDG': 96, 'PHX': 96,
    'FAT': 72, 'MOD': 72, 'LAS': 72, 'PLM': 72, 'YUM': 72,
    'DEN': 120, 'PDX': 120, 'SEA': 120,
    'HNL': 168
}
df['Time_Limit_Hrs'] = df['目的中心'].map(time_limits)
df['Overdue'] = (df['包裹状态'] == '已妥投') & (df['Total_Duration_Hrs'] > df['Time_Limit_Hrs'])

# ========== 超时包裹明细 ==========
overdue_df = df[df['Overdue']].copy()
overdue_df['超时时间（小时）'] = overdue_df['Total_Duration_Hrs'] - overdue_df['Time_Limit_Hrs']
overdue_detail = overdue_df[[
    '运单号', '目的中心', '包裹状态', '妥投时间', '统计日期', 
    'Total_Duration_Hrs', 'Time_Limit_Hrs', '超时时间（小时）'
]].rename(columns={
    '运单号': 'Tracking #',
    '目的中心': 'Center',
    '包裹状态': 'Status',
    '妥投时间': 'Delivered Time',
    '统计日期': 'Date',
    'Total_Duration_Hrs': 'Actual Duration (hrs)',
    'Time_Limit_Hrs': 'SLA (hrs)',
    '超时时间（小时）': 'Overdue (hrs)'
})

# ========== 显示超时数据 ==========
st.subheader("⏰ Overdue Packages Summary")
st.metric("超时包裹数", len(overdue_detail))
st.dataframe(overdue_detail, use_container_width=True)

# ========== 下载按钮 ==========
download_buf = io.BytesIO()
with pd.ExcelWriter(download_buf, engine="openpyxl") as writer:
    overdue_detail.to_excel(writer, index=False, sheet_name="Overdue Packages")

st.download_button(
    label="📥 Download Overdue Package Detail as Excel",
    data=download_buf.getvalue(),
    file_name="overdue_packages.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
